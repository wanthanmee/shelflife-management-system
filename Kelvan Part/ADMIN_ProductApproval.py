from customtkinter import *
import tkinter as tk
from tkinter import ttk, messagebox
import tkinter.font as tkfont
from tkinter import ttk
import sqlite3
import re
import barcode
from barcode.writer import ImageWriter
import os
from datetime import datetime
from PIL import Image, ImageTk
from tkcalendar import DateEntry
import openpyxl
from fpdf import FPDF  # pip install fpdf

'''
Additional Features: 
+Filter by Registration Date
+Filter by Owner Name in Search Bar
+Barcode Generation
'''
DB_NAME = "ProductRegistration.db"

def generate_barcode(product_data):
    try:
        # Create a unique identifier using product data
        # Combine ID, batch_id, and timestamp to ensure uniqueness
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        barcode_data = f"{product_data[0]}{product_data[1]}{timestamp}"
        
        # Generate barcode
        barcode_instance = barcode.get('code128', barcode_data, writer=ImageWriter())
        
        # Create barcodes directory if it doesn't exist
        if not os.path.exists('barcodes'):
            os.makedirs('barcodes')
            
        # Save barcode as PNG
        barcode_path = f"barcodes/{barcode_data}.png"
        barcode_instance.save(f"barcodes/{barcode_data}")
        
        return barcode_data, barcode_path
    except Exception as e:
        raise Exception(f"Failed to generate barcode: {str(e)}")

class ProductDetailPage(CTkToplevel):
    def __init__(self, master, product_data):
        super().__init__(master)
        self.product_data = product_data  # Convert to list for easier modification
        self.title("Product Details")
        self.geometry("1000x900")

        # Make window modal
        self.transient(master)
        self.grab_set()

        # Dictionary to store entry widgets
        self.entry_widgets = {}
        self.build_ui()

    def build_ui(self):
        # Main container frame
        main_frame = CTkFrame(self)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Title
        title_label = CTkLabel(main_frame, text="Product Approval", font=("Arial", 24))
        title_label.pack(pady=10)

        # Create two columns: details and barcode
        content_frame = CTkFrame(main_frame)
        content_frame.pack(fill="both", expand=True, pady=10)

        # Left column for details
        details_frame = CTkFrame(content_frame)
        details_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Scrollable frame for details
        canvas = tk.Canvas(details_frame)
        scrollbar = ttk.Scrollbar(details_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = CTkFrame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Product details with edit functionality
        labels = ["ID:", "Batch ID:", "Product Name:", "Description:", "Submission Date:", 
                 "Testing Date:", "Maturity Date:", "Test Completed:", "Test ID:", 
                 "Test Result Location:", "Date Updated:", "Updated By:", "Owner ID:", 
                 "Status:", "Barcode:"]
        
        editable_fields = ["batch_id", "product_name", "description", "submission_date", 
                          "testing_date", "maturity_date", "test_completed", "test_id", 
                          "test_result_location", "date_updated", "updated_by", "owner_id"]

        for i, (label, value) in enumerate(zip(labels, self.product_data)):
            # Label
            CTkLabel(scrollable_frame, text=label, font=("Arial", 14, "bold")).grid(
                row=i, column=0, sticky="w", pady=(10, 0), padx=10)
            
            # Value or Entry field
            field_name = label.lower().replace(":", "").replace(" ", "_")
            if field_name in editable_fields:
                # Create a frame for the entry and edit button
                field_frame = CTkFrame(scrollable_frame)
                field_frame.grid(row=i, column=1, sticky="w", padx=10, pady=(10, 0))
                
                # Entry widget
                entry = CTkEntry(field_frame, width=300)
                entry.insert(0, str(value))
                entry.pack(side="left", padx=(0, 5))
                
                # Edit button
                edit_btn = CTkButton(
                    field_frame,
                    text="Edit",
                    width=60,
                    command=lambda e=entry, l=label: self.save_edit(e, l)
                )
                edit_btn.pack(side="left")
                
                # Store the entry widget
                self.entry_widgets[label] = entry
            else:
                # Non-editable field
                value_label = CTkLabel(scrollable_frame, text=str(value), font=("Arial", 12))
                value_label.grid(row=i, column=1, sticky="w", padx=10, pady=(10, 0))

                # Save reference to status label
                if label == "Status:":
                    self.status_label = value_label

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Right column for barcode
        self.barcode_frame = CTkFrame(content_frame)
        self.barcode_frame.pack(side="right", fill="both", padx=(10, 0))

        # Display barcode if it exists
        if self.product_data[14]:  # If barcode exists
            try:
                barcode_path = f"barcodes/{self.product_data[14]}.png"
                if os.path.exists(barcode_path):
                    barcode_image = Image.open(barcode_path)
                    max_width = 300
                    ratio = max_width / barcode_image.width
                    new_size = (max_width, int(barcode_image.height * ratio))
                    barcode_image = barcode_image.resize(new_size, Image.Resampling.LANCZOS)
                    
                    barcode_photo = ImageTk.PhotoImage(barcode_image)
                    
                    barcode_label = CTkLabel(self.barcode_frame, image=barcode_photo, text="")
                    barcode_label.image = barcode_photo
                    barcode_label.pack(pady=10)
                    
                    CTkLabel(self.barcode_frame, text=f"Barcode: {self.product_data[14]}", 
                            font=("Arial", 12)).pack(pady=5)
            except Exception as e:
                print(f"Error displaying barcode: {e}")

        # Buttons frame at the bottom
        buttons_frame = CTkFrame(main_frame)
        buttons_frame.pack(pady=20, fill="x")

        # Approve button
        approve_button = CTkButton(
            buttons_frame,
            text="Approve",
            command=self.approve_owner,
            fg_color="green",
            hover_color="dark green"
        )
        approve_button.pack(side="left", padx=10, expand=True)

        # Deny button
        deny_button = CTkButton(
            buttons_frame,
            text="Deny",
            command=self.deny_owner,
            fg_color="red",
            hover_color="dark red"
        )
        deny_button.pack(side="left", padx=10, expand=True)

        # Close button
        close_button = CTkButton(
            buttons_frame,
            text="Close",
            command=self.destroy,
            fg_color="gray",
            hover_color="dark gray"
        )
        close_button.pack(side="right", padx=10, expand=True)

    def save_edit(self, entry_widget, label):
        try:
            # Get the new value
            new_value = entry_widget.get()
            
            # Get the field name from the label
            field_name = label.lower().replace(":", "").replace(" ", "_")
            
            # Update the database
            conn = sqlite3.connect('ProductRegistration.db')
            cursor = conn.cursor()
            
            # Map the field names to database column names
            field_mapping = {
                "batch_id": "batch_id",
                "product_name": "product_name",
                "description": "description",
                "submission_date": "submission_date",
                "testing_date": "testing_date",
                "maturity_date": "maturity_date",
                "test_completed": "test_completed",
                "test_id": "test_id",
                "test_result_location": "test_result_location",
                "date_updated": "date_updated",
                "updated_by": "updated_by",
                "owner_id": "owner_id"
            }
            
            if field_name in field_mapping:
                db_field = field_mapping[field_name]
                # Update the specific field
                cursor.execute(f"UPDATE products SET {db_field} = ? WHERE id = ?",
                             (new_value, self.product_data[0]))
                
                conn.commit()
                conn.close()
                
                # Update the product_data list
                field_index = {
                    "batch_id": 1, "product_name": 2, "description": 3,
                    "submission_date": 4, "testing_date": 5, "maturity_date": 6,
                    "test_completed": 7, "test_id": 8, "test_result_location": 9,
                    "date_updated": 10, "updated_by": 11, "owner_id": 12
                }
                
                if field_name in field_index:
                    self.product_data[field_index[field_name]] = new_value
                
                # Show success message
                messagebox.showinfo("Success", f"{label} updated successfully!")
                
                # Refresh the main window's data
                if hasattr(self.master, 'load_data'):
                    self.master.load_data()
            else:
                messagebox.showerror("Error", f"Field {label} is not editable")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update {label}: {str(e)}")
            print(f"Error details: {str(e)}")  # For debugging

    def refresh_barcode_display(self):
        # Clear barcode frame
        for widget in self.barcode_frame.winfo_children():
            widget.destroy()

        barcode_data = self.product_data[14] if len(self.product_data) > 14 else None
        barcode_path = self.product_data[15] if len(self.product_data) > 15 else None

        if barcode_data and barcode_path and os.path.exists(barcode_path):
            try:
                barcode_image = Image.open(barcode_path)
                max_width = 300
                ratio = max_width / barcode_image.width
                new_size = (max_width, int(barcode_image.height * ratio))
                barcode_image = barcode_image.resize(new_size, Image.Resampling.LANCZOS)
                barcode_photo = ImageTk.PhotoImage(barcode_image)

                barcode_label = CTkLabel(self.barcode_frame, image=barcode_photo, text="")
                barcode_label.image = barcode_photo  # Keep reference!
                barcode_label.pack(pady=10)

                CTkLabel(self.barcode_frame, text=f"Barcode: {barcode_data}",
                        font=("Arial", 12)).pack(pady=5)

            except Exception as e:
                print(f"Error displaying barcode after refresh: {e}")

    def update_status(self, new_status):
        try:
            product_id = self.product_data[0]

            # Generate barcode first if status is being approved
            if new_status == "Approved" and self.product_data[13] != "Approved":
                try:
                    # Generate barcode (must not touch DB inside this function!)
                    barcode_data, barcode_path = generate_barcode(self.product_data)

                    # Now update everything inside one database transaction
                    with sqlite3.connect("ProductRegistration.db", timeout=10) as conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            "UPDATE products SET status = ?, barcode_data = ?, barcode_path = ? WHERE id = ?",
                            (new_status, barcode_data, barcode_path, product_id)
                        )

                    # Update local product_data list
                    self.product_data[13] = new_status
                    self.product_data[14] = barcode_data
                    if len(self.product_data) > 15:
                        self.product_data[15] = barcode_path
                    else:
                        self.product_data.append(barcode_path)

                    messagebox.showinfo("Success",
                        f"Product approved and barcode generated!\n"
                        f"Barcode: {barcode_data}\nSaved to: {barcode_path}")

                    self.refresh_barcode_display()

                except Exception as e:
                    messagebox.showerror("Barcode Error", f"Failed to generate barcode:\n{e}")
                    return

            else:
                # Just update status
                with sqlite3.connect("ProductRegistration.db", timeout=10) as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE products SET status = ? WHERE id = ?",
                        (new_status, product_id)
                    )
                self.product_data[13] = new_status

            # Update the status label in the UI
            if hasattr(self, "status_label"):
                self.status_label.configure(text=new_status)

                # Optionally color it based on status
                if new_status == "Approved":
                    self.status_label.configure(text_color="green")
                elif new_status == "Rejected":
                    self.status_label.configure(text_color="red")
                else:
                    self.status_label.configure(text_color="black")

        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to update status:\n{e}")

    def approve_owner(self):
        if messagebox.askyesno("Confirm Approval", "Are you sure you want to approve this product?"):
            self.update_status("Approved")

    def deny_owner(self):
        if messagebox.askyesno("Confirm Denial", "Are you sure you want to deny this product?"):
            self.update_status("Denied")
    '''
    =====================================================================================================
    Function: def load_data(self)
    Description: 
    This function loads data from the SQLite database into the treeview.
    It applies filters based on the search term and status filter.
    =====================================================================================================
    '''
    def load_data(self):
        try:
            conn = sqlite3.connect('ProductRegistration.db')
            cursor = conn.cursor()

            query = '''
            SELECT id, batch_id, product_name, description, submission_date, testing_date, maturity_date, test_completed, test_id, test_result_location, date_updated, updated_by, owner_id, status, barcode
            FROM products
            WHERE 1=1
            '''
            params = []

            # Search filter
            search_term = self.search_var.get().strip()
            if search_term:
                query += " AND (product_name LIKE ?)"
                like_term = f"%{search_term}%"
                params += [like_term]

            # Status filter
            selected_status = self.status_filter_var.get()
            if selected_status in ["Pending", "Approved", "Denied"]:
                query += " AND status = ?"
                params.append(selected_status)

            # Date filter
            submission_date = self.submission_date_var.get().strip()
            if submission_date and re.match(r"^\d{4}-\d{2}$", submission_date):
                query += " AND strftime('%Y-%m', submission_date) = ?"
                params.append(submission_date)

            cursor.execute(query, params)
            rows = cursor.fetchall()

            # Clear old data
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Insert new rows
            for row in rows:
                self.tree.insert("", "end", values=row)

            conn.close()
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to load data:\n{e}")

    '''
    =====================================================================================================
    Function: def on_row_double_click(self, event)
    Description: 
    This function handles the double-click event on a row in the treeview.
    It retrieves the data of the selected row and prints it.
    =====================================================================================================
    '''
    def on_treeview_double_click(self, event):
        selected_item = self.tree.focus()  # or self.tree.selection()[0]
        if selected_item:
            values = self.tree.item(selected_item)["values"]
            if values:
                    product_data = list(values)
                    ProductDetailPage(self, product_data)

    '''
    =====================================================================================================
    Function: def open_detail_popup(self, data)
    Description: 
    This function opens a popup window to display and edit the details of a selected product.
    It allows the user to modify product details and save changes to both the SQLite database and an Excel file.
    =====================================================================================================
    '''
    def open_detail_popup(self, data):
        detail_popup = CTkToplevel(self)
        detail_popup.title("Product Details")
        detail_popup.geometry("600x400")

        fields = ["Product Name", "Description", "Testing Date", "Maturity Date"]
        entries = {}

        for i, label in enumerate(fields):
            CTkLabel(detail_popup, text=label).pack()
            entry = CTkEntry(detail_popup)
            entry.insert(0, data[i+1])  # Skip Batch_ID at data[0]
            entry.pack()
            entries[label] = entry
        '''
        =====================================================================================================
        Function: def save_changes()
        Description: 
        This function saves the changes made in the detail popup to both the SQLite database and an Excel file.
        It retrieves the values from the entry fields, updates the database, and modifies the Excel file accordingly.
        =====================================================================================================
        '''
        def save_changes():
            product_name = entries["Product Name"].get()
            description = entries["Description"].get()
            testing_date = entries["Testing Date"].get()
            maturity_date = entries["Maturity Date"].get()
            batch_id = data[0]  # Assuming Batch_ID is the first value

            # Update SQLite
            try:
                conn = sqlite3.connect(DB_NAME)
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE products 
                    SET product_name = ?, description = ?, testing_date = ?, maturity_date = ? 
                    WHERE batch_id = ?
                """, (product_name, description, testing_date, maturity_date, batch_id))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("SQLite Error", str(e))
                return

            # Update Excel
            try:
                wb = openpyxl.load_workbook("ProductRecords.xlsx")
                ws = wb.active
                for row in ws.iter_rows(min_row=2):  # Assuming headers in row 1
                    if row[0].value == batch_id:
                        row[1].value = product_name
                        row[2].value = description
                        row[4].value = testing_date
                        row[5].value = maturity_date
                        break
                wb.save("ProductRecords.xlsx")
            except Exception as e:
                messagebox.showerror("Excel Error", str(e))
                return

            messagebox.showinfo("Success", "Changes saved to database and Excel.")
            detail_popup.destroy()
            self.load_data()

        CTkButton(detail_popup, text="Save Changes", command=save_changes).pack(pady=10)
    '''
    =====================================================================================================
    Function: def export_to_excel(self)
    Description: 
    This function exports the data from the treeview to an Excel file.
    It creates a new workbook, adds the column headers, and appends the data from the treeview.
    =====================================================================================================
    '''
    def export_to_excel(self):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(self.tree["columns"])
            for row in self.tree.get_children():
                ws.append(self.tree.item(row)["values"])
            wb.save("ProductRecords.xlsx")
            messagebox.showinfo("Success", "Exported to ProductRecords.xlsx")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))
    '''
    =====================================================================================================
    Function: def export_to_pdf(self)
    Description: 
    This function exports the data from the treeview to a PDF file.
    It uses the FPDF library to create a PDF document, adds a page, sets the font, and writes the data.
    =====================================================================================================
    '''
    def export_to_pdf(self):
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=10)
            for col in self.tree["columns"]:
                pdf.cell(25, 10, col[:15], border=1)
            pdf.ln()
            for row in self.tree.get_children():
                for val in self.tree.item(row)["values"]:
                    pdf.cell(25, 10, str(val)[:15], border=1)
                pdf.ln()
            pdf.output("ProductRecords.pdf")
            messagebox.showinfo("Success", "Exported to ProductRecords.pdf")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

class ProductListPage(CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.pack(fill="both", expand=True)
        self.build_ui()

    def build_ui(self):
        # Search + Filter Frame
        filter_frame = CTkFrame(self)
        filter_frame.pack(pady=10, fill="x")

        CTkLabel(filter_frame, text="Search: ").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *_: self.load_data())
        CTkEntry(filter_frame, textvariable=self.search_var, placeholder_text="Search by Product Name").pack(side="left", fill="x", expand=True, padx=10)

        self.status_filter = tk.StringVar(value="All")
        CTkOptionMenu(filter_frame, values=["All", "Pending", "Approved", "Denied"], variable=self.status_filter, command=lambda x: self.load_data()).pack(side="left", padx=5)

        # Treeview for products
        tree_frame = CTkFrame(self)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(tree_frame, show="headings", columns=(
            "id", "batch_id", "product_name", "description", "submission_date",
            "testing_date", "maturity_date", "test_completed", "test_id",
            "test_result_location", "date_updated", "updated_by", "owner_id", "status", "barcode"
        ))
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center")

        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self.open_detail_popup)

        self.load_data()

    def load_data(self):
        try:
            conn = sqlite3.connect("ProductRegistration.db")
            cursor = conn.cursor()

            query = "SELECT * FROM products WHERE 1=1"
            params = []

            if self.search_var.get().strip():
                query += " AND product_name LIKE ?"
                params.append(f"%{self.search_var.get().strip()}%")

            if self.status_filter.get() != "All":
                query += " AND status = ?"
                params.append(self.status_filter.get())

            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()

            for item in self.tree.get_children():
                self.tree.delete(item)

            for row in rows:
                self.tree.insert("", "end", values=row)
        except Exception as e:
            messagebox.showerror("Database Error", f"Could not load data: {e}")

    def open_detail_popup(self, event):
        selected_item = self.tree.focus()
        if selected_item:
            product_data = self.tree.item(selected_item)["values"]
            if product_data:
                ProductDetailPage(self, product_data)  # Open the popup

'''
=====================================================================================================
Run the app.
This is the main entry point of the application.
It sets the appearance mode and default color theme, creates the main window, and starts the application.
=====================================================================================================
'''
if __name__ == "__main__":
    app = CTk()
    app.geometry("1200x800")
    set_appearance_mode("light")
    app.title("Admin - Product Approval Dashboard")

    ProductListPage(app)  # Show the list page

    app.mainloop()
   

