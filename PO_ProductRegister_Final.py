from customtkinter import *
import tkinter as tk
from PIL import Image, ImageTk
from customtkinter import CTkFont
from tkinter import messagebox
from datetime import datetime
from tkcalendar import DateEntry
import openpyxl
from openpyxl import Workbook
import os
import uuid
import sqlite3

file_name = "ProductRecords.xlsx"

class ProductRegistration(CTkFrame):
    '''
    =====================================================================================================
    Constructor: def __init__(self, parent, controller=None):
    Description: 
    This function initializes the ProductRegistration frame, which is a subclass of CTkFrame.
    It sets up the frame to fill the parent window and configures its appearance.
    =====================================================================================================
    '''
    def __init__(self, parent, controller=None, owner_id=None):
        super().__init__(parent)
        self.controller = controller
        self.owner_id = owner_id

                # Configure this frame
        self.pack(fill="both", expand=True)
        self.configure(fg_color="white")

        # Variables
        self.productName_var = tk.StringVar()
        self.description_var = tk.StringVar()
        self.testingDate_var = tk.StringVar()
        self.maturityDate_var = tk.StringVar()

        self.build_ui()
    '''
    =====================================================================================================
    Function: def build_ui(self)
    Description: 
    This function builds the UI for the Product List page.
    It includes a title, search bar, status filter, and a treeview to display product data.
    =====================================================================================================
    '''
    def build_ui(self):
        # Title
        title_label = CTkLabel(self, text="Product Register", font=("Arial", 30), text_color="#654633")
        title_label.pack(pady=50)

        # Product Frame
        self.product_frame = CTkFrame(self, fg_color="#FCEBEB", width=1000, height=300, corner_radius=20)
        self.product_frame.pack(pady=10, padx=20, fill="x")

        CTkLabel(self.product_frame, text="Product Details", font=("Arial", 24), text_color="#5B3E2B").place(x=20, y=30)
        CTkLabel(self.product_frame, text="Product Name:", font=("Arial", 16)).place(x=20, y=80)
        CTkEntry(self.product_frame, width=800, height=30, textvariable=self.productName_var).place(x=150, y=80)

        CTkLabel(self.product_frame, text="Description:", font=("Arial", 16)).place(x=20, y=140)
        CTkEntry(self.product_frame, width=800, height=100, textvariable=self.description_var).place(x=150, y=140)

        # Date Frame
        self.date_frame = CTkFrame(self, fg_color="#F0EFF8", width=1000, height=250, corner_radius=20)
        self.date_frame.pack(pady=10, padx=20, fill="x")

        CTkLabel(self.date_frame, text="Date Selection", font=("Arial", 24), text_color="#5B3E2B").place(x=20, y=30)

        CTkLabel(self.date_frame, text="Testing Date:", font=("Arial", 16)).place(x=20, y=80)
        self.testing_date_entry = DateEntry(self.date_frame, width=18, background='pink', foreground='white',
                                            borderwidth=2, textvariable=self.testingDate_var)
        self.testing_date_entry.place(x=150, y=80)

        CTkLabel(self.date_frame, text="Maturity Date:", font=("Arial", 16)).place(x=20, y=130)
        self.maturity_date_entry = DateEntry(self.date_frame, width=18, background='pink', foreground='white',
                                             borderwidth=2, textvariable=self.maturityDate_var)
        self.maturity_date_entry.place(x=150, y=130)

        # Buttons
        CTkButton(self.date_frame, text="Save Product", command=self.save_data, font=("Arial", 18),
                  fg_color="#95d194", hover_color="#FDC09A").place(x=650, y=180)
        CTkButton(self.date_frame, text="Clear Form", command=self.clear_form, font=("Arial", 18),
                  fg_color="#f16c6c", hover_color="#FDC09A").place(x=800, y=180)
    '''
    =====================================================================================================
    Function: def save_data(self)
    Description: 
    This function saves the product data entered in the form to a SQLite database and an Excel file.
    It checks for required fields, creates a database table if it doesn't exist, and inserts the data.
    =====================================================================================================
    '''
    def save_data(self):
        productName = self.productName_var.get()
        description = self.description_var.get()
        try:
            testingDate = datetime.strptime(self.testingDate_var.get(), "%m/%d/%y").strftime("%Y-%m-%d")
            maturityDate = datetime.strptime(self.maturityDate_var.get(), "%m/%d/%y").strftime("%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Date Format Error", "Please enter valid dates.")
            return

        if not productName or not description or not testingDate or not maturityDate:
            messagebox.showerror("Error", "All fields are required!")
            return

        submissionDate = datetime.now().strftime("%Y-%m-%d")
        test_completed = "No"
        test_id = str(uuid.uuid4())[:8]
        date_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_by = "System"
        owner_id = self.owner_id 

        try:
            conn = sqlite3.connect('ProductRegistration.db')
            cursor = conn.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    batch_id TEXT UNIQUE,
                    product_name TEXT,
                    description TEXT,
                    submission_date TEXT,
                    testing_date TEXT,
                    maturity_date TEXT,
                    test_completed TEXT,
                    test_id TEXT,
                    test_result_location TEXT,
                    date_updated TEXT,
                    updated_by TEXT,
                    owner_id TEXT  -- NEW COLUMN
                )
            ''')

            cursor.execute('''
                INSERT INTO products (
                    product_name, description, submission_date,
                    testing_date, maturity_date, test_completed, test_id,
                    test_result_location, date_updated, updated_by, owner_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                productName, description, submissionDate,
                testingDate, maturityDate, test_completed, test_id,
                "", date_updated, updated_by, owner_id
            ))

            last_id = cursor.lastrowid
            batch_id = str(last_id).zfill(5)

            base_directory = "TestResults"
            batch_directory = os.path.join(base_directory, submissionDate, batch_id)
            os.makedirs(batch_directory, exist_ok=True)
            testResultLocation = os.path.join(batch_directory, f"{productName}_results.txt")

            cursor.execute('''
                UPDATE products
                SET batch_id = ?, test_result_location = ?
                WHERE id = ?
            ''', (batch_id, testResultLocation, last_id))

            conn.commit()
            conn.close()

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"An error occurred while saving to database:\n{e}")
            return

        # Save to Excel
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            headers = ["Batch ID", "Product Name", "Description", "Submission Date", "Testing Date",
                    "Maturity Date", "Test Completed", "Test ID", "Test Result Location", "Date Updated", "Updated By", "Owner ID"]
            ws.append(headers)

            for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                for cell in col:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="FFD3B5", end_color="FFD3B5", fill_type="solid")
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        else:
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active

        ws.append([
            batch_id, productName, description, submissionDate, testingDate,
            maturityDate, test_completed, test_id, testResultLocation, date_updated, updated_by, owner_id
        ])

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        wb.save(file_name)
        messagebox.showinfo("Success", f"Data saved!\nResults stored at:\n{testResultLocation}")
        self.clear_form()
    '''
    =====================================================================================================
    Function: def clear_form(self)
    Description: 
    This function clears all the input fields in the product registration form.
    It resets the StringVar variables to empty strings.
    =====================================================================================================
    '''
    def clear_form(self):
        self.productName_var.set("")
        self.description_var.set("")
        self.testingDate_var.set("")
        self.maturityDate_var.set("")

if __name__ == "__main__":
    app = CTk()
    app.geometry("1200x800")
    set_appearance_mode("light")
    app.title("Product Registration")

    product_page = ProductRegistration(app)
    app.mainloop()
