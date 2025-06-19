from customtkinter import *
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry  # pip install tkcalendar
import tkinter.font as tkfont
from tkinter import ttk
import sqlite3
import openpyxl
from fpdf import FPDF  # pip install fpdf

'''
Additional Features: 
+Filter by Submission Date
+Filter by Owner Name in Search Bar
'''
DB_NAME = "ProductRegistration.db"

class ProductListPage(CTkFrame):
    '''
    =====================================================================================================
    Constructor: def __init__(self, parent, controller=None):
    Description: 
    This function initializes the Product List page.
    It sets up the main frame and calls the build_ui function to create the UI elements.
    =====================================================================================================
    '''
    def __init__(self, parent, controller=None, owner_id=None):
        super().__init__(parent)   # Use parent here, not master
        self.controller = controller
        self.owner_id = owner_id

        print(f"Owner ID: {self.owner_id}")  # Debugging line to check owner_id

        self.configure(fg_color="white")
        self.build_ui()
    '''
    =====================================================================================================
    Function: def build_ui(self)
    Description: 
    This function builds the UI for the Product List page.
    It includes a title, search bar, status filter, and a treeview to display product data.
    =====================================================================================================
    '''
    def build_ui(self):

        # Title
        title_label = CTkLabel(self, text="Product List", font=("Arial", 30),
                               text_color="#654633")
        title_label.pack(pady=50)
        
        '''
        =====================================================================================================
        SET UP STYLE FOR TREEVIEW
        =====================================================================================================
        '''
        # Setup bold heading font style
        style = ttk.Style()
        style.theme_use("clam")  # Switch to a theme that respects styling
        
        # Set the font for the Treeview headings
        bold_heading_font = tkfont.Font(family="Arial", size=12, weight="bold")
        style.configure("Treeview.Heading", font=bold_heading_font)

        # Configure Treeview heading style
        style.configure("Treeview.Heading",
        font= bold_heading_font,
            foreground="#654633",  # Text color
            background="#FEDEE9")  # Background color

        style.map("Treeview",
          background=[("selected", "#F8E5E5")],  # Light blue selection
          foreground=[("selected", "#654633")])  # Dark text
        
        # Search bar
        search_frame = CTkFrame(self, fg_color="white")
        search_frame.pack(pady=5, padx=20, fill="x")

        # Search label
        search_label = CTkLabel(search_frame, text="Search: ", font=("Arial", 16))
        search_label.pack(side="left")

        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *args: self.load_data()) #calls the load_data function automatically

        self.search_entry = CTkEntry(
        search_frame,
        placeholder_text="Search Product Name",
        textvariable=self.search_var
        )
        self.search_entry.pack(side="left", padx=(0, 10), fill="x", expand=True)
        self.search_entry.bind("<KeyRelease>", lambda event: self.load_data())

        # Status filter
        self.status_filter_var = tk.StringVar(value="All")
        self.status_dropdown = CTkOptionMenu(search_frame, values=["All", "Pending", "Approved", "Rejected"], 
                                             variable=self.status_filter_var, 
                                             font=("Arial", 16),
                                             fg_color="#654633",
                                             text_color="white",
                                             command=lambda x: self.load_data())
        self.status_dropdown.pack(side="left")

        # Submission date filter
        date_filter_frame = CTkFrame(self, fg_color="white")
        date_filter_frame.pack(pady=5, padx=20, fill="x")

        CTkLabel(date_filter_frame, text="From:", font=("Arial", 16)).pack(side="left")
        self.start_date = DateEntry(date_filter_frame, width=12)
        self.start_date.pack(side="left", padx=5)

        CTkLabel(date_filter_frame, text="To:", font=("Arial", 16)).pack(side="left")
        self.end_date = DateEntry(date_filter_frame, width=12)
        self.end_date.pack(side="left", padx=5)

        filter_button = CTkButton(date_filter_frame, text="Apply Date Filter", font=("Arial", 15), 
                                 fg_color="#654633", 
                                 hover_color="#FDC09A",
                                 text_color="white", 
                                 command=self.load_data)
        filter_button.pack(side="left", padx=10)

        # Treeview Frame
        tree_frame = CTkFrame(self)
        tree_frame.pack(padx=20, pady=10, fill="both", expand=True)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview
        self.tree = ttk.Treeview(
            tree_frame,
            columns=("Batch_ID", "Product Name", "Description", "Submission Date", "Testing Date", "Maturity Date", "Test Completed", "Test_ID", "Test Result Location", "Date Updated", "Updated By"),
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set
        )
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        # Layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center")

        # Select Record
        self.tree.bind("<Double-1>", self.on_row_double_click)

        # Load data
        self.load_data()

        button_frame = CTkFrame(self, fg_color="white")
        button_frame.pack(pady=10)

        export_excel_btn = CTkButton(button_frame, text="EXCEL", font=("Arial", 18),
                                     fg_color="#95d194", hover_color="#FDC09A",
                                     command=self.export_to_excel)
        export_excel_btn.pack(side="left", padx=10)

        export_pdf_btn = CTkButton(button_frame, text="PDF", font=("Arial", 18),
                                   fg_color="#f16c6c", hover_color="#FDC09A",
                                    command=self.export_to_pdf)
        export_pdf_btn.pack(side="left", padx=10)

    '''
    =====================================================================================================
    Function: def load_data(self)
    Description: 
    This function loads data from the SQLite database into the treeview.
    It applies filters based on the search term and status filter.
    =====================================================================================================
    '''
    def load_data(self):
        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()

            query = '''
            SELECT batch_id, product_name, description, submission_date,
                testing_date, maturity_date, test_completed, test_id,
                test_result_location, date_updated, updated_by, owner_id, status
            FROM products
            WHERE owner_id = ?  -- Only get data for the logged-in Product Owner
            '''
            params = [self.owner_id]  # ← Filter by current Product Owner

            # Search filter
            search_term = self.search_var.get().strip()
            if search_term:
                query += " AND product_name LIKE ?"
                like_term = f"%{search_term}%"
                params.append(like_term)

            # Status filter
            selected_status = self.status_filter_var.get()
            if selected_status in ["Pending", "Approved", "Rejected"]:
                query += " AND status = ?"
                params.append(selected_status)
            
            # Submission Date filter
            start = self.start_date.get_date()
            end = self.end_date.get_date()
            query += " AND date(submission_date) BETWEEN ? AND ?"
            params += [start, end]

            cursor.execute(query, params)
            rows = cursor.fetchall()

            # Clear old data
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Insert new rows
            for row in rows:
                self.tree.insert("", "end", values=row)

            conn.close()
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to load data:\n{e}")
    '''
    =====================================================================================================
    Function: def on_row_double_click(self, event)
    Description: 
    This function handles the double-click event on a row in the treeview.
    It retrieves the data of the selected row and prints it.
    =====================================================================================================
    '''
    def on_row_double_click(self, event):
        selected_item = self.tree.focus()
        if selected_item:
            row_data = self.tree.item(selected_item, "values")
            self.open_detail_popup(row_data)

    '''
    =====================================================================================================
    Function: def open_detail_popup(self, data)
    Description: 
    This function opens a popup window to display and edit the details of a selected product.
    It allows the user to modify product details and save changes to both the SQLite database and an Excel file.
    =====================================================================================================
    '''
    def open_detail_popup(self, data):
        detail_popup = CTkToplevel(self)
        detail_popup.title("Product Details")
        detail_popup.geometry("600x400")

        fields = ["Product Name", "Description", "Testing Date", "Maturity Date"]
        entries = {}

        for i, label in enumerate(fields):
            CTkLabel(detail_popup, text=label).pack()
            entry = CTkEntry(detail_popup)
            entry.insert(0, data[i+1])  # Skip Batch_ID at data[0]
            entry.pack()
            entries[label] = entry
        '''
        =====================================================================================================
        Function: def save_changes()
        Description: 
        This function saves the changes made in the detail popup to both the SQLite database and an Excel file.
        It retrieves the values from the entry fields, updates the database, and modifies the Excel file accordingly.
        =====================================================================================================
        '''
        def save_changes():
            product_name = entries["Product Name"].get()
            description = entries["Description"].get()
            testing_date = entries["Testing Date"].get()
            maturity_date = entries["Maturity Date"].get()
            batch_id = data[0]  # Assuming Batch_ID is the first value

            # Update SQLite
            try:
                conn = sqlite3.connect(DB_NAME)
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE products 
                    SET product_name = ?, description = ?, testing_date = ?, maturity_date = ? 
                    WHERE batch_id = ?
                """, (product_name, description, testing_date, maturity_date, batch_id))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("SQLite Error", str(e))
                return

            # Update Excel
            try:
                wb = openpyxl.load_workbook("ProductRecords.xlsx")
                ws = wb.active
                for row in ws.iter_rows(min_row=2):  # Assuming headers in row 1
                    if row[0].value == batch_id:
                        row[1].value = product_name
                        row[2].value = description
                        row[4].value = testing_date
                        row[5].value = maturity_date
                        break
                wb.save("ProductRecords.xlsx")
            except Exception as e:
                messagebox.showerror("Excel Error", str(e))
                return

            messagebox.showinfo("Success", "Changes saved to database and Excel.")
            detail_popup.destroy()
            self.load_data()

        CTkButton(detail_popup, text="Save Changes", command=save_changes).pack(pady=10)
    '''
    =====================================================================================================
    Function: def export_to_excel(self)
    Description: 
    This function exports the data from the treeview to an Excel file.
    It creates a new workbook, adds the column headers, and appends the data from the treeview.
    =====================================================================================================
    '''
    def export_to_excel(self):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(self.tree["columns"])
            for row in self.tree.get_children():
                ws.append(self.tree.item(row)["values"])
            wb.save("ProductRecords.xlsx")
            messagebox.showinfo("Success", "Exported to ProductRecords.xlsx")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))
    '''
    =====================================================================================================
    Function: def export_to_pdf(self)
    Description: 
    This function exports the data from the treeview to a PDF file.
    It uses the FPDF library to create a PDF document, adds a page, sets the font, and writes the data.
    =====================================================================================================
    '''
    def export_to_pdf(self):
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=10)
            for col in self.tree["columns"]:
                pdf.cell(25, 10, col[:15], border=1)
            pdf.ln()
            for row in self.tree.get_children():
                for val in self.tree.item(row)["values"]:
                    pdf.cell(25, 10, str(val)[:15], border=1)
                pdf.ln()
            pdf.output("ProductRecords.pdf")
            messagebox.showinfo("Success", "Exported to ProductRecords.pdf")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))


# Run standalone if executed directly
if __name__ == "__main__":
    app = CTk()
    app.geometry("800x800")
    set_appearance_mode("light")
    app.title("Product List")

    product_page = ProductListPage(app)
    product_page.pack(fill="both", expand=True)

    app.mainloop()

